import openpyxl
import os
import sys
import time
import datetime
import random
import calendar



str_path = "柔韧性提升训练数据"

path_file = ""


def get_files():
    global path_file

    for each_file in os.listdir(os.curdir):
        if  str_path in each_file:
            path_file = each_file


if __name__ == '__main__':

    # print("备注：")

    get_files()

    if path_file == "":
        input("同文件夹下未找到两个对应表格文件模板，请检查文件位置或文件命名")
        sys.exit()

    print("开始进行"+str_path+"的处理")

    wb_path = openpyxl.load_workbook(str(path_file))
    path_sheet = wb_path.worksheets[0]
    data_sheet = wb_path.worksheets[1]

    for x in range(2, 101):  # 增加数据
        #获取此学生坐位体前屈成绩
        zuowei = path_sheet.cell(row=x, column=8).value

        zhanwei = 0

        #性别读取
        gender = 1
        if path_sheet.cell(row=x, column=2).value ==2:
            gender =2

        #计算对应于体感站位的成绩
        if  gender==1: #男
            zhanwei =round( -0.9588*zuowei +31.371 +random.uniform(-2,2),1)
        else:#女
            zhanwei = round(-1*zuowei +30.8+random.uniform(-2,2) ,1)

        #写入站位成绩
        path_sheet.cell(row=x, column=11, value=zhanwei)

        #查询站位成绩所在等级
        zhanwei_rank = 11
        for j in range(5,25):
            if gender ==1:
                if zhanwei <= data_sheet.cell(row=j, column=5).value:
                    zhanwei_rank = data_sheet.cell(row=j, column=2).value
                    break
            else:
                if zhanwei <= data_sheet.cell(row=j, column=8).value:
                    zhanwei_rank = data_sheet.cell(row=j, column=2).value
                    break
        
        #填充量化等级
        path_sheet.cell(row=x, column=12, value=zhanwei_rank)

        #训练一个月后的成绩
        perfor_af_1mth =0
        if path_sheet.cell(row=x, column=10).value == "优秀":
            perfor_af_1mth = round(zuowei + random.uniform(-0.5,2),1)
        elif path_sheet.cell(row=x, column=10).value == "良好":
            perfor_af_1mth = round(zuowei + random.uniform(-0.5,3),1)
        elif path_sheet.cell(row=x, column=10).value == "及格":
            perfor_af_1mth = round(zuowei + random.uniform(-0.5,4),1)
        else:
            perfor_af_1mth = round(zuowei + random.uniform(-0.5,5),1)

        path_sheet.cell(row=x, column=13, value=perfor_af_1mth)

        #训练两个月后的成绩
        perfor_af_2mth =0
        if path_sheet.cell(row=x, column=10).value == "优秀":
            perfor_af_2mth = round(perfor_af_1mth + random.uniform(0.5,2),1)
        elif path_sheet.cell(row=x, column=10).value == "良好":
            perfor_af_2mth = round(perfor_af_1mth + random.uniform(0.5,3),1)
        elif path_sheet.cell(row=x, column=10).value == "及格":
            perfor_af_2mth = round(perfor_af_1mth + random.uniform(0.5,4),1)
        else:
            perfor_af_2mth = round(perfor_af_1mth + random.uniform(0.5,5),1)

        path_sheet.cell(row=x, column=14, value=perfor_af_2mth)

    print(str_path+"生成中，请稍后。。。")

    wb_path.save(path_file)


    print("生成完毕")

    input("请按任意键退出")
