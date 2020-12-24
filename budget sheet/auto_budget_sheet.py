# 本项目用于项目申报通过已有申报书的预算表和更改后的总预算自动按比例生成新的预算表
# 项目预算模板按照省重大专项布置

import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os


list_dec_name = ["购置设备费", "自制设备费", "设备改造与租赁", "材料费", "测试化验加工费",
                 "燃料动力费", "信息费（出版/文献/信息传播/知识产权事物费）",
                 "差旅费/会议费/国际合作与交流费", "专家咨询费", "劳务费", "其他支出"]

list_indec_name = ["管理费", "绩效支出", "其他"]

# 用于调整列表数据值，调整至目标和


def ajust_num(listnum, tar_sum):
    list_int_num = [round(exp) for exp in listnum]
    while np.sum(list_int_num) != tar_sum:
        index = list_int_num.index(max(list_int_num))

        if (np.sum(list_int_num) > tar_sum):
            list_int_num[index] -= 1
        else:
            list_int_num[index] += 1

    return list_int_num


# 完成一次输入和按总款项比例计算
def cocu_data():

    # 参考文件的费用明细
    list_ref_dec_expense = []
    list_ref_indec_expense = []

    ref_SUM = 0  # 参考文件总费用

    while 1:
        print("一、直接费用")
        print("  设备费")

        for i in range(3):
            temp = input("    "+list_dec_name[i]+"：")
            if temp == '':
                temp = '0'
            list_ref_dec_expense.append(int(temp))

        for i in range(3, 11):
            temp = input("  "+list_dec_name[i]+"：")
            if temp == '':
                temp = '0'
            list_ref_dec_expense.append(int(temp))

        print("二、间接费用")
        for i in range(3):
            temp = input("  "+list_indec_name[i]+"：")
            if temp == '':
                temp = '0'
            list_ref_indec_expense.append(int(temp))

        ref_equip_expense = (list_ref_dec_expense[0] +
                             list_ref_dec_expense[1]+list_ref_dec_expense[2])
        ref_dec_expense = np.sum(list_ref_dec_expense)
        ref_indec_expense = np.sum(list_ref_indec_expense)
        ref_SUM = ref_dec_expense + ref_indec_expense

        print("设备费合计：", ref_equip_expense, "，直接费用合计：",
              ref_dec_expense, "，间接费用合计：", ref_indec_expense, "，总计：", ref_SUM)

        check = input("请检查计算结果是否正确？若正确直接回车，不正确则输入“1”回车重新输入")

        if check != "1":
            break
        else:
            list_ref_dec_expense = []
            list_ref_indec_expense = []

    budget = int(input("请输入当前待分配的总预算数目："))
    print("正在按比例生成各个经费预算内容。。。。")

    ratio = round(budget/ref_SUM, 2)

    list_dec_expense_f = [exp * ratio for exp in list_ref_dec_expense]
    list_indec_expense_f = [exp * ratio for exp in list_ref_indec_expense]

    # 圆整数据
    list_int_expense = ajust_num(
        list_dec_expense_f + list_indec_expense_f, budget)
    print(list_int_expense)

    print("本项目根据比例生成的预算结果：")
    print("  一、直接费用")
    print("    1、设备费")
    for i in range(3):
        print("      ("+str(i)+"、)" +
              list_dec_name[i]+"：", str(list_int_expense[i]))

    for i in range(3, 11):
        print("    "+str(i)+"、"+list_dec_name[i]+"：", str(list_int_expense[i]))

    print("  二、间接费用")
    for i in range(3):
        print("    "+str(i)+"、" +
              list_indec_name[i]+"：", str(list_int_expense[11+i]))

    return list_int_expense


def create_xml(list_all_budget, list_gov_budget):
    print("正在生成excel文件，请稍后。。。")
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active

    if(list_gov_budget == []):
        list_gov_budget = [0 for x in range(len(list_all_budget))]

    # 创建表头
    ws['A1'] = "资金支出预算"  # write
    ws['B1'] = "预算金额"  # write
    ws['C1'] = "其中：省财政拨款"  # write

    # 创建各行
    ws['A2'] = "一、直接费用"  # write
    ws['B2'] = "=SUM(B4:B14)"  # 公式
    ws['C2'] = "=SUM(C4:C14)"  # 公式

    ws['A3'] = "1、设备费"  # write
    ws['B3'] = "=SUM(B4:B6)"  # 公式
    ws['C3'] = "=SUM(C4:C6)"  # 公式

    for i in range(3):
        ws.cell(row=i+4, column=1, value="("+str(i+1)+")" + list_dec_name[i])
        ws.cell(row=i+4, column=2, value=list_all_budget[i])
        ws.cell(row=i+4, column=3, value=list_gov_budget[i])

    for i in range(3, 11):
        ws.cell(row=i+4, column=1, value=str(i-1)+"、" + list_dec_name[i])
        ws.cell(row=i+4, column=2, value=list_all_budget[i])
        ws.cell(row=i+4, column=3, value=list_gov_budget[i])

    ws['A15'] = " 二、间接费用"
    ws['B15'] = "=SUM(B16:B18)"  # 公式
    ws['C15'] = "=SUM(C16:C18)"  # 公式

    for i in range(3):
        ws.cell(row=i+16, column=1, value=str(i+1)+"、" + list_indec_name[i])
        ws.cell(row=i+16, column=2, value=list_all_budget[i+11])
        ws.cell(row=i+16, column=3, value=list_gov_budget[i+11])

    ws['A19'] = " 支出合计"
    ws['B19'] = "=SUM(B4:B15)"  # 公式
    ws['C19'] = "=SUM(C4:C15)"  # 公式

    # 设置行高
    for i in range(1, 20):
        ws.row_dimensions[i].height = 20

    # 设置列宽
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17

    # 设置居中
    for rowOfCellObjects in ws['A1':'C19']:
        for cellObj in rowOfCellObjects:
            cellObj.alignment = Alignment(horizontal='left', vertical='center')

        #ws['A1':'C19'].alignment = Alignment(horizontal='left', vertical='center')

    wb.save("budget.xlsx")

    print("当前文件目录位于：", os.getcwd(), "\\budget.xlsx")  # 获取当前路径

    input("生成成功，公式已添加，如需修改请在excel文件基础上微调，按回车关闭并打开excel文件")

    os.startfile("budget.xlsx")


if __name__ == '__main__':

    current_work_dir = os.path.dirname(__file__)  # 当前文件所在的目录
    os.chdir(current_work_dir)

    # os.startfile("budget.xlsx")

    print("请输入参照文件的项目经费预算表内容（单位：万元）：")
    list_all_budget = cocu_data()

    list_gov_budget = []
    temp = input("是否继续生成财政拨款部分？默认继续，输入0不考虑此部分")
    if temp != '0':
        print("请输入参照文件的财政拨款预算表内容（单位：万元）：")
        list_gov_budget = cocu_data()

    temp = input("是否生成结果excel文件？默认继续，输入0退出")
    if temp != '0':
        create_xml(list_all_budget, list_gov_budget)
