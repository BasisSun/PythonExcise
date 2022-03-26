
import openpyxl
import os
import sys
import time
import datetime
import random
import calendar


str_health = "每日身体健康监测记录表"
str_path = "幼儿行动轨迹追踪表"

path_date_rownum = 6  # 轨迹表的日期行是第六行
health_date_rownum = 5  # 健康表的日期行是第五行

path_date_colnum = 3  # 轨迹表的每月1日所在列是第3列
health_date_colnum = 4  # 健康表的每月1日所在列是第4列

health_file = ""
path_file = ""


def get_files():
    global health_file
    global path_file

    for each_file in os.listdir(os.curdir):
        if date0 in each_file:
            if str_health in each_file:
                health_file = each_file
            elif str_path in each_file:
                path_file = each_file


def str_to_list(abs_day):
    list_days = abs_day.split(',')
    list_days_to_move = []

    for day0 in list_days:
        if '-' in day0:
            day_part = day0.split('-')
            day_part_fill = range(int(day_part[0]), int(day_part[1])+1)

            day_part_fill = [str(x) for x in day_part_fill]  # 转化为字符串

            list_days_to_move.append(day0)

            list_days = list_days+day_part_fill

    if len(list_days_to_move) > 0:
        for daytomove in list_days_to_move:
            list_days.remove(daytomove)

    return list_days


if __name__ == '__main__':

    print("备注：")
    print("1.请将程序放入两张表格模板同一目录下；")
    print("2.请将两张表格分别按：\n"+"XX班"+str_health+"202009\n" +
          "XX班"+str_path+"202009\n的文件名命名，年月改成对应时间")
    print("3.模板需要提供31日的日期空间\n")

    date0 = input("请输入表格年月（如：202009）；")
    str_year = date0[0:4]
    str_month = date0[4:]

    get_files()

    if path_file == "" or health_file == "":
        input("同文件夹下未找到两个对应表格文件模板，请检查文件位置或文件命名")
        sys.exit()

    monthRange = calendar.monthrange(int(str_year), int(str_month))  # 获取当月日期天数
    print(monthRange[1])

    holidays = input("请输入本月长假期,用逗号隔开（如：1-7,20-23）；")
    list_holidays = str_to_list(holidays)

    list_casual_leave = []  # 事假记录

    print("开始进行"+str_path+"的处理")

    wb_path = openpyxl.load_workbook(str(path_file))
    path_sheet = wb_path.worksheets[0]

    for x in range(3, 40):  # 修改表格月份和天数
        if "月份" in str(path_sheet.cell(row=path_date_rownum-2, column=x).value):
            path_sheet.cell(row=path_date_rownum-2, column=x +
                            1, value=str_year+"年"+str_month+"月")
        if x <= monthRange[1] + 2:
            path_sheet.cell(row=path_date_rownum, column=x, value=str(x-2)+"日")
        elif path_sheet.cell(row=path_date_rownum, column=x).value != None and "日" in path_sheet.cell(row=path_date_rownum, column=x).value:
            path_sheet.cell(row=path_date_rownum, column=x, value="")
            row_i = path_date_rownum + 1
            while path_sheet.cell(row=row_i, column=1).value == row_i-path_date_rownum:
                path_sheet.cell(row=row_i, column=x, value="")
                row_i += 1

    i = path_date_rownum + 1  # 第一个学生从此行开始

    print("请输入下面学生的缺席日期，用英文逗号隔开，并回车（例如：9,11,15-18）,输入 ch 将撤销并回到上一次输入")

    while path_sheet.cell(row=i, column=1).value == i-path_date_rownum:

        hitString = str(path_sheet.cell(row=i, column=1).value) + \
            ""+str(path_sheet.cell(row=i, column=2).value)+"："
        print(hitString)

        abs_day = input("    请输入病假日期：")

        if abs_day == "ch" and i > path_date_rownum + 1:
            i = i-1
            list_casual_leave.pop()
            continue

        list_abs_day = str_to_list(abs_day)

        str_casual_leave_day = input("    请输入事假日期：")
        if str_casual_leave_day == "ch" and i > path_date_rownum + 1:
            i = i-1
            list_casual_leave.pop()
            continue

        casual_leave_day = str_to_list(str_casual_leave_day)
        list_casual_leave.append(casual_leave_day)

        list_abs_day = list_abs_day + casual_leave_day

        j = path_date_colnum  # 每月1日所在列

        while str(path_sheet.cell(row=path_date_rownum, column=j).value) == str(j-2)+"日":
            anyday = datetime.datetime(int(str_year), int(
                str_month), j-2).strftime("%w")  # 判断当天是周几
            if anyday == '6' or anyday == '0' or str(j-2) in list_holidays:
                path_sheet.cell(row=i, column=j, value="√")
            elif str(j-2) in list_abs_day:
                path_sheet.cell(row=i, column=j, value="√")
            else:
                path_sheet.cell(row=i, column=j, value="FL8:10-4:40无异常")

            j = j+1

        i = i + 1

    print(str_path+"生成中，请稍后。。。")

    wb_path.save(path_file)

    print("开始进行"+str_health+"的处理")

    wb_health = openpyxl.load_workbook(str(health_file))
    health_sheet = wb_health.worksheets[0]

    for x in range(4, 40):
        if "月份" in str(health_sheet.cell(row=health_date_rownum-2, column=x).value):
            health_sheet.cell(row=health_date_rownum-2,
                              column=x+2, value=str_year+"年"+str_month+"月")
        if x <= monthRange[1] + 3:
            health_sheet.cell(row=health_date_rownum,
                              column=x, value=str(x-3)+"日")
        elif health_sheet.cell(row=health_date_rownum, column=x).value != None and "日" in health_sheet.cell(row=health_date_rownum, column=x).value:
            health_sheet.cell(row=health_date_rownum, column=x, value="")
            row_i = health_date_rownum + 1
            while health_sheet.cell(row=row_i, column=1).value == row_i-health_date_rownum:
                health_sheet.cell(row=row_i, column=x, value="")
                row_i += 1

    # 再次遍历轨迹表
    i = path_date_rownum + 1  # 第一个学生从此行开始

    while path_sheet.cell(row=i, column=1).value == i-path_date_rownum:
        j = path_date_colnum  # 每月1日所在列
        while str(path_sheet.cell(row=path_date_rownum, column=j).value) == str(j-2)+"日":

            if path_sheet.cell(row=i, column=j).value == "√":
                count_true = 0
                begin_true = j

                valid_sec = 1  # 用于记录是否该区间段为纯病假期放假

                while path_sheet.cell(row=i, column=j).value == "√":
                    anyday = datetime.datetime(int(str_year), int(
                        str_month), j-2).strftime("%w")  # 判断当天是周几
                    if anyday == '6' or anyday == '0' or str(j-2) in list_holidays or str(j-2) in list_casual_leave[i-path_date_rownum-1]:
                        health_sheet.cell(row=i - path_date_rownum + health_date_rownum,
                                          column=j-path_date_colnum+health_date_colnum, value="√")
                    else:
                        valid_sec = 0

                    count_true += 1
                    j += 1
                else:
                    if str(path_sheet.cell(row=path_date_rownum, column=j).value) == str(j-2)+"日":
                        health_sheet.cell(row=i - path_date_rownum + health_date_rownum,
                                          column=j-path_date_colnum+health_date_colnum, value="√")

                if valid_sec == 0:
                    if count_true > 2:
                        for k in range(count_true):
                            health_sheet.cell(row=i - path_date_rownum + health_date_rownum,
                                              column=begin_true+k-path_date_colnum+health_date_colnum, value="D")
                    elif count_true == 2:
                        for k in range(count_true):
                            health_sheet.cell(row=i - path_date_rownum + health_date_rownum,
                                              column=begin_true+k-path_date_colnum+health_date_colnum, value="B")
                    else:
                        health_sheet.cell(row=i - path_date_rownum + health_date_rownum, column=begin_true -
                                          path_date_colnum+health_date_colnum, value=random.choice(["C", "E"]))

            else:
                health_sheet.cell(row=i - path_date_rownum + health_date_rownum,
                                  column=j-path_date_colnum+health_date_colnum, value="√")

            j += 1

        i += 1

    print(str_path, "正在生成。。。")
    wb_health.save(health_file)

    print("生成完毕")

    input("请按任意键退出")
